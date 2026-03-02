"""
Canvas Results Aggregation Service
===================================
Fetches quiz submissions and enrollment grades from Canvas,
computes statistics, and prepares data for export.
"""
import io
import logging
import math
from statistics import mean, median, stdev

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Optional

from backend.services import canvas_service

logger = logging.getLogger(__name__)


# ============================================================================
# Quiz-level results
# ============================================================================

async def get_quiz_results(
    token: str,
    base_url: str,
    course_id: int,
    quiz_id: int,
) -> dict:
    """
    Fetch all quiz submissions and compute aggregated statistics.
    Returns a ``QuizResultsAggregation``-style dict.
    """

    # 1. Fetch quiz metadata
    quizzes_res = await canvas_service.list_quizzes(token, base_url, course_id)
    quiz_title = "Unknown Quiz"
    points_possible: float | None = None
    if quizzes_res["success"]:
        for q in quizzes_res.get("quizzes", []):
            if q.get("id") == quiz_id:
                quiz_title = q.get("title", quiz_title)
                points_possible = q.get("points_possible")
                break

    # 2. Fetch quiz submissions
    subs_res = await canvas_service.get_quiz_submissions(
        token, base_url, course_id, quiz_id
    )
    if not subs_res["success"]:
        return {
            "success": False,
            "error": subs_res.get("error"),
            "quiz_id": quiz_id,
            "quiz_title": quiz_title,
        }

    raw_subs = subs_res["quiz_submissions"]

    # 3. Transform submissions
    submissions: list[dict] = []
    scores: list[float] = []
    graded_count = 0

    for s in raw_subs:
        sub = {
            "user_id": s.get("user_id"),
            "user_name": s.get("user_id"),  # placeholder — name not in quiz_submissions endpoint
            "submission_id": s.get("id"),
            "attempt": s.get("attempt"),
            "score": s.get("score"),
            "kept_score": s.get("kept_score"),
            "points_possible": s.get("quiz_points_possible"),
            "started_at": s.get("started_at"),
            "finished_at": s.get("finished_at"),
            "workflow_state": s.get("workflow_state", "unknown"),
        }
        submissions.append(sub)

        if sub["workflow_state"] == "complete" and sub["kept_score"] is not None:
            scores.append(float(sub["kept_score"]))
            graded_count += 1

    # 4. Compute statistics
    stats: dict = {
        "average_score": None,
        "median_score": None,
        "max_score": None,
        "min_score": None,
        "std_dev": None,
    }
    if scores:
        stats["average_score"] = round(mean(scores), 2)
        stats["median_score"] = round(median(scores), 2)
        stats["max_score"] = max(scores)
        stats["min_score"] = min(scores)
        if len(scores) > 1:
            stats["std_dev"] = round(stdev(scores), 2)

    # 5. Score distribution (buckets of 10%)
    distribution: dict[str, int] = {}
    if scores and points_possible and points_possible > 0:
        buckets = 10
        for sc in scores:
            pct = (sc / points_possible) * 100
            bucket_idx = min(int(pct // 10), buckets - 1)
            low = bucket_idx * 10
            high = low + 10
            label = f"{low}-{high}%"
            distribution[label] = distribution.get(label, 0) + 1

    return {
        "success": True,
        "quiz_id": quiz_id,
        "quiz_title": quiz_title,
        "points_possible": points_possible,
        "total_submissions": len(submissions),
        "graded_count": graded_count,
        **stats,
        "score_distribution": distribution,
        "submissions": submissions,
    }


# ============================================================================
# Course-level enrollment grades
# ============================================================================

async def get_course_grades(
    token: str,
    base_url: str,
    course_id: int,
) -> dict:
    """
    Fetch all student enrollments and compute grade aggregations.
    Returns a ``CourseGradesAggregation``-style dict.
    """
    enroll_res = await canvas_service.get_course_enrollments(
        token, base_url, course_id, enrollment_type="StudentEnrollment"
    )
    if not enroll_res["success"]:
        return {
            "success": False,
            "error": enroll_res.get("error"),
            "course_id": course_id,
        }

    enrollments_raw = enroll_res["enrollments"]

    enrollments: list[dict] = []
    current_scores: list[float] = []
    final_scores: list[float] = []

    for e in enrollments_raw:
        grades = e.get("grades", {})
        item = {
            "user_id": e.get("user_id"),
            "user_name": e.get("user", {}).get("name"),
            "enrollment_id": e.get("id"),
            "enrollment_state": e.get("enrollment_state", "unknown"),
            "current_score": grades.get("current_score"),
            "final_score": grades.get("final_score"),
            "current_grade": grades.get("current_grade"),
            "final_grade": grades.get("final_grade"),
        }
        enrollments.append(item)

        if item["current_score"] is not None:
            current_scores.append(float(item["current_score"]))
        if item["final_score"] is not None:
            final_scores.append(float(item["final_score"]))

    # Stats
    avg_current = round(mean(current_scores), 2) if current_scores else None
    avg_final = round(mean(final_scores), 2) if final_scores else None
    max_current = max(current_scores) if current_scores else None
    min_current = min(current_scores) if current_scores else None

    # Grade letter distribution (if available)
    grade_dist: dict[str, int] = {}
    for e in enrollments:
        grade = e.get("current_grade") or "N/A"
        grade_dist[grade] = grade_dist.get(grade, 0) + 1

    # Course name — try to get from enrollments (Canvas includes course info)
    course_name = None
    if enrollments_raw:
        course_name = enrollments_raw[0].get("course_section_name")

    return {
        "success": True,
        "course_id": course_id,
        "course_name": course_name,
        "total_students": len(enrollments),
        "average_current_score": avg_current,
        "average_final_score": avg_final,
        "max_current_score": max_current,
        "min_current_score": min_current,
        "grade_distribution": grade_dist,
        "enrollments": enrollments,
    }


# ============================================================================
# Export helpers
# ============================================================================

async def export_quiz_results_csv(
    token: str,
    base_url: str,
    course_id: int,
    quiz_id: int,
) -> tuple[str, str]:
    """
    Export quiz results as CSV string.
    Returns (csv_content, filename).
    """
    data = await get_quiz_results(token, base_url, course_id, quiz_id)
    if not data.get("success"):
        raise ValueError(data.get("error", "Failed to fetch results"))

    lines = ["user_id,submission_id,attempt,score,kept_score,points_possible,workflow_state,started_at,finished_at"]
    for s in data["submissions"]:
        line = ",".join(str(s.get(k, "")) for k in [
            "user_id", "submission_id", "attempt", "score", "kept_score",
            "points_possible", "workflow_state", "started_at", "finished_at"
        ])
        lines.append(line)

    filename = f"quiz_{quiz_id}_results.csv"
    return "\n".join(lines), filename


async def export_course_grades_csv(
    token: str,
    base_url: str,
    course_id: int,
) -> tuple[str, str]:
    """
    Export course enrollments / grades as CSV string.
    Returns (csv_content, filename).
    """
    data = await get_course_grades(token, base_url, course_id)
    if not data.get("success"):
        raise ValueError(data.get("error", "Failed to fetch grades"))

    lines = ["user_id,user_name,enrollment_id,enrollment_state,current_score,final_score,current_grade,final_grade"]
    for e in data["enrollments"]:
        name = str(e.get("user_name", "")).replace(",", " ")
        line = ",".join([
            str(e.get("user_id", "")),
            name,
            str(e.get("enrollment_id", "")),
            str(e.get("enrollment_state", "")),
            str(e.get("current_score", "")),
            str(e.get("final_score", "")),
            str(e.get("current_grade", "")),
            str(e.get("final_grade", "")),
        ])
        lines.append(line)

    filename = f"course_{course_id}_grades.csv"
    return "\n".join(lines), filename


# ============================================================================
# Excel export helpers
# ============================================================================

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, col_count: int):
    """Apply header styling to row 1."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


async def export_quiz_results_excel(
    token: str,
    base_url: str,
    course_id: int,
    quiz_id: int,
) -> tuple[bytes, str]:
    """
    Export quiz results as Excel (.xlsx) bytes.
    Returns (xlsx_bytes, filename).
    """
    data = await get_quiz_results(token, base_url, course_id, quiz_id)
    if not data.get("success"):
        raise ValueError(data.get("error", "Failed to fetch results"))

    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = ["Metric", "Value"]
    ws_summary.append(summary_headers)
    _style_header_row(ws_summary, len(summary_headers))

    ws_summary.append(["Quiz ID", data["quiz_id"]])
    ws_summary.append(["Quiz Title", data["quiz_title"]])
    ws_summary.append(["Points Possible", data.get("points_possible")])
    ws_summary.append(["Total Submissions", data["total_submissions"]])
    ws_summary.append(["Graded Count", data["graded_count"]])
    ws_summary.append(["Average Score", data.get("average_score")])
    ws_summary.append(["Median Score", data.get("median_score")])
    ws_summary.append(["Max Score", data.get("max_score")])
    ws_summary.append(["Min Score", data.get("min_score")])
    ws_summary.append(["Std Dev", data.get("std_dev")])
    _auto_width(ws_summary)

    # ---------- Sheet 2: Submissions ----------
    ws_subs = wb.create_sheet("Submissions")
    sub_headers = [
        "User ID", "Submission ID", "Attempt", "Score",
        "Kept Score", "Points Possible", "% Score",
        "Workflow State", "Started At", "Finished At",
    ]
    ws_subs.append(sub_headers)
    _style_header_row(ws_subs, len(sub_headers))

    pp = data.get("points_possible") or 0
    for s in data["submissions"]:
        score = s.get("score")
        pct = round((score / pp) * 100, 1) if score is not None and pp > 0 else None
        ws_subs.append([
            s.get("user_id"),
            s.get("submission_id"),
            s.get("attempt"),
            score,
            s.get("kept_score"),
            s.get("points_possible"),
            pct,
            s.get("workflow_state", ""),
            s.get("started_at", ""),
            s.get("finished_at", ""),
        ])

    # Color-code % column
    pct_fill_hi = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pct_fill_mid = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    pct_fill_lo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in range(2, ws_subs.max_row + 1):
        cell = ws_subs.cell(row=row, column=7)  # % Score column
        if cell.value is not None:
            if cell.value >= 80:
                cell.fill = pct_fill_hi
            elif cell.value >= 50:
                cell.fill = pct_fill_mid
            else:
                cell.fill = pct_fill_lo

    _auto_width(ws_subs)

    # ---------- Sheet 3: Score Distribution ----------
    dist = data.get("score_distribution", {})
    if dist:
        ws_dist = wb.create_sheet("Distribution")
        ws_dist.append(["Range", "Count"])
        _style_header_row(ws_dist, 2)
        for label, count in sorted(dist.items()):
            ws_dist.append([label, count])
        _auto_width(ws_dist)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"quiz_{quiz_id}_results.xlsx"
    return buf.getvalue(), filename


async def export_course_grades_excel(
    token: str,
    base_url: str,
    course_id: int,
) -> tuple[bytes, str]:
    """
    Export course enrollment grades as Excel (.xlsx) bytes.
    Returns (xlsx_bytes, filename).
    """
    data = await get_course_grades(token, base_url, course_id)
    if not data.get("success"):
        raise ValueError(data.get("error", "Failed to fetch grades"))

    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = ["Metric", "Value"]
    ws_summary.append(summary_headers)
    _style_header_row(ws_summary, len(summary_headers))

    ws_summary.append(["Course ID", data["course_id"]])
    ws_summary.append(["Course Name", data.get("course_name")])
    ws_summary.append(["Total Students", data["total_students"]])
    ws_summary.append(["Average Current Score", data.get("average_current_score")])
    ws_summary.append(["Average Final Score", data.get("average_final_score")])
    ws_summary.append(["Max Current Score", data.get("max_current_score")])
    ws_summary.append(["Min Current Score", data.get("min_current_score")])
    _auto_width(ws_summary)

    # ---------- Sheet 2: Enrollments ----------
    ws_enroll = wb.create_sheet("Enrollments")
    enroll_headers = [
        "User ID", "User Name", "Enrollment ID", "Enrollment State",
        "Current Score", "Final Score", "Current Grade", "Final Grade",
    ]
    ws_enroll.append(enroll_headers)
    _style_header_row(ws_enroll, len(enroll_headers))

    pct_fill_hi = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pct_fill_mid = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    pct_fill_lo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for e in data["enrollments"]:
        ws_enroll.append([
            e.get("user_id"),
            e.get("user_name", ""),
            e.get("enrollment_id"),
            e.get("enrollment_state", ""),
            e.get("current_score"),
            e.get("final_score"),
            e.get("current_grade", ""),
            e.get("final_grade", ""),
        ])

    # Color-code current_score column
    for row in range(2, ws_enroll.max_row + 1):
        cell = ws_enroll.cell(row=row, column=5)  # Current Score
        if cell.value is not None:
            if cell.value >= 80:
                cell.fill = pct_fill_hi
            elif cell.value >= 50:
                cell.fill = pct_fill_mid
            else:
                cell.fill = pct_fill_lo
    _auto_width(ws_enroll)

    # ---------- Sheet 3: Grade Distribution ----------
    g_dist = data.get("grade_distribution", {})
    if g_dist:
        ws_dist = wb.create_sheet("Grade Distribution")
        ws_dist.append(["Grade", "Count"])
        _style_header_row(ws_dist, 2)
        for grade, count in sorted(g_dist.items()):
            ws_dist.append([grade, count])
        _auto_width(ws_dist)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"course_{course_id}_grades.xlsx"
    return buf.getvalue(), filename
